""" DCF Model """
__docformat__ = "numpy"

from urllib.request import urlopen
from typing import List, Union, Dict, Any, Tuple
from zipfile import ZipFile
from io import BytesIO
import re

from sklearn.linear_model import LinearRegression
from openpyxl import worksheet
from bs4 import BeautifulSoup
import financedatabase as fd
import yfinance as yf
import pandas as pd
import requests

from gamestonk_terminal.stocks.fundamental_analysis import dcf_static


def string_float(string: str) -> float:
    """Convert a string to a float

    Parameters
    ----------
    string : str
        String to be converted

    Returns
    -------
    number : float
        Analysis of filings text
    """
    if string.strip().replace(",", "").replace("-", "") == "":
        return 0
    return float(string.strip().replace(",", "").replace("-", ""))


def insert_row(
    name: str, index: str, df: pd.DataFrame, row_v: List[str]
) -> pd.DataFrame:
    """Allows a row to be added given an index and name

    Parameters
    ----------
    name : str
        Name to be added to df
    index : str
        The row the new item will go after
    df : pd.DataFrame
        The dataframe to be modified
    row_v : List[str]
        The items to be added to the row

    Returns
    -------
    new_df : pd.DataFrame
        The new dataframe
    """
    pd.options.mode.chained_assignment = None
    if name not in df.index:
        row_number = df.index.get_loc(index) + 1
        df1 = df[0:row_number]
        df2 = df[row_number:]
        df1.loc[name] = row_v
        df_result = pd.concat([df1, df2])
        return df_result
    return df


def set_cell(
    ws: worksheet,
    cell: str,
    text: Union[int, str, float] = None,
    font: str = None,
    border: str = None,
    fill: str = None,
    alignment: str = None,
    num_form: str = None,
):
    """Set the value for a cell

    Parameters
    ----------
    ws : worksheet
        The worksheet to be modified
    cell : str
        The cell that will be modified
    text : Union[int, str, float]
        The new value of the cell
    font : str
        The type of font
    border : str
        The type of border
    fill : str
        The type of fill
    alignment : str
        The type of alignment
    num_form : str
        The format for numbers
    """
    if text:
        ws[cell] = text
    if font:
        ws[cell].font = font
    if border:
        ws[cell].border = border
    if fill:
        ws[cell].fill = fill
    if alignment:
        ws[cell].alignment = alignment
    if num_form:
        ws[cell].number_format = num_form


def get_fama_raw() -> pd.DataFrame:
    """Get Fama French data

    Returns
    -------
    df : pd.DataFrame
        Fama French data
    """
    with urlopen(
        "http://mba.tuck.dartmouth.edu/pages/faculty/ken.french/ftp/F-F_Research_Data_Factors_CSV.zip"
    ) as url:

        # Download Zipfile and create pandas DataFrame
        with ZipFile(BytesIO(url.read())) as zipfile:
            with zipfile.open("F-F_Research_Data_Factors.CSV") as zip_open:
                df = pd.read_csv(
                    zip_open,
                    header=0,
                    names=["Date", "MKT-RF", "SMB", "HML", "RF"],
                    skiprows=3,
                )

    df = df[df["Date"].apply(lambda x: len(str(x).strip()) == 6)]
    df["Date"] = df["Date"].astype(str) + "01"
    df["Date"] = pd.to_datetime(df["Date"], format="%Y%m%d")
    df["MKT-RF"] = pd.to_numeric(df["MKT-RF"], downcast="float")
    df["SMB"] = pd.to_numeric(df["SMB"], downcast="float")
    df["HML"] = pd.to_numeric(df["HML"], downcast="float")
    df["RF"] = pd.to_numeric(df["RF"], downcast="float")
    df["MKT-RF"] = df["MKT-RF"] / 100
    df["SMB"] = df["SMB"] / 100
    df["HML"] = df["HML"] / 100
    df["RF"] = df["RF"] / 100
    df = df.set_index("Date")
    return df


def get_historical_5(ticker: str) -> pd.DataFrame:
    """Get 5 year monthly historical performance for a ticker with dividends filtered

    Parameters
    ----------
    ticker : str
        The ticker to be analyzed

    Returns
    -------
    df : pd.DataFrame
        Historical data
    """
    tick = yf.Ticker(ticker)
    df = tick.history(period="5y", interval="1mo")
    df = df[df.index.to_series().apply(lambda x: x.day == 1)]
    df = df.drop(["Dividends", "Stock Splits"], axis=1)
    df = df.dropna()
    return df


def get_fama_coe(ticker: str) -> float:
    """Use Fama and French to get the cost of equity for a company

    Parameters
    ----------
    ticker : str
        The ticker to be analyzed

    Returns
    -------
    coef : float
        The stock's Fama French coefficient
    """
    df_f = get_fama_raw()
    df_h = get_historical_5(ticker)
    df = df_h.join(df_f)
    df = df.dropna()
    df["Monthly Return"] = df["Close"].pct_change()
    df["Excess Monthly Return"] = df["Monthly Return"] - df["RF"]
    df = df.dropna()
    x = df[["MKT-RF", "SMB", "HML"]]
    y = df["Excess Monthly Return"]

    model = LinearRegression().fit(x, y)
    coefs = model.coef_
    return (
        df["RF"].mean()
        + coefs[0] * df["MKT-RF"].mean()
        + coefs[1] * df["SMB"].mean()
        + coefs[2] * df["HML"].mean()
    ) * 12


def others_in_sector(ticker: str, sector: str, industry: str) -> List[str]:
    """Get other stocks in a ticker's sector

    Parameters
    ----------
    ticker : str
        The ticker to be excluded
    sector : str
        The sector to pull from
    industry : str
        The industry to pull from

    Returns
    -------
    tickers : List[str]
        List of tickers in the same sector
    """
    industry = industry.replace("—", " - ")
    industry = industry.replace("/", " ")

    sister_ticks = fd.select_equities(sector=sector, industry=industry)

    # This filters sisters to match market cap and removes ticker analyzed
    if ticker in sister_ticks:
        market_cap = sister_ticks[ticker]["market_cap"]
        sister_ticks.pop(ticker, None)
        sister_ticks = {
            k: v for (k, v) in sister_ticks.items() if v["market_cap"] == market_cap
        }
    sister_ticks = list(sister_ticks)
    return sister_ticks


def create_dataframe(ticker: str, statement: str):
    """
    Creates a df financial statement for a given ticker

    Parameters
    ----------
    ticker : str
        The ticker to create a dataframe for
    statement : str
        The financial statement dataframe to create

    Returns
    -------
    statement : pd.DataFrame
        The financial statement requested
    rounding : int
        The amount of rounding to use
    """
    URL = f"https://stockanalysis.com/stocks/{ticker}/financials/"
    URL += dcf_static.statement_url[statement]
    ignores = dcf_static.statement_ignore[statement]

    r = requests.get(URL, headers=dcf_static.headers)

    if "404 - Page Not Found" in r.text:
        return pd.DataFrame(), None
    soup = BeautifulSoup(r.content, "html.parser")

    table = soup.find("table", attrs={"class": re.compile("fintbl")})
    if table is None:
        return pd.DataFrame(), None
    head = table.find("thead")
    if head is None:
        return pd.DataFrame(), None
    columns = head.find_all("th")
    if columns is None:
        return pd.DataFrame(), None

    years = [x.get_text().strip() for x in columns if "-" not in x.get_text().strip()]
    len_data = len(years) - 1

    phrase = soup.find("div", attrs={"class": "text-sm text-gray-600 block lg:hidden"})
    phrase = phrase.get_text().lower()

    if "thousand" in phrase:
        rounding = 1_000
    elif "millions" in phrase:
        rounding = 1_000_000
    elif "billions" in phrase:
        rounding = 1_000_000_000
    else:
        return pd.DataFrame(), None

    body = table.find("tbody")
    rows = body.find_all("tr")

    all_data = [
        [
            x.get_text().strip() if x.get_text().strip() != "-" else "0"
            for x in y.find_all("td")
        ]
        for y in rows
    ]

    df = pd.DataFrame(data=all_data)
    df = df.loc[:, ~(df == "Upgrade").any()]
    df = df.set_index(0)
    n = df.shape[1] - len_data
    if n > 0:
        df = df.iloc[:, :-n]
    df.columns = years[1 : len(df.columns) + 1]

    for ignore in ignores:
        if ignore in df.index:
            df = df.drop([ignore])
    df = df[df.columns[::-1]]

    if statement == "IS":
        vals = ["Revenue", dcf_static.gaap_is]
    elif statement == "BS":
        vals = ["Cash & Equivalents", dcf_static.gaap_bs]
    elif statement == "CF":
        vals = ["Net Income", dcf_static.gaap_cf]

    if vals[0] in df.index:
        blank_list = ["0" for _ in df.loc[vals[0]].to_list()]
    else:
        return pd.DataFrame(), None
    for i, _ in enumerate(vals[1][1:]):
        df = insert_row(vals[1][i + 1], vals[1][i], df, blank_list)

    return df, rounding


def get_sister_dfs(ticker: str, info: Dict[str, Any], n: int):
    """
    Get dataframes for sister companies

    Parameters
    ----------
    ticker : str
        The ticker to create a dataframe for
    into : Dict[str,Any]
        The dictionary produced from the yfinance.info function
    n : int
        The number of sister companies to produce

    Returns
    -------
    new_list: List[str, pd.DataFrame]
        A list of sister companies
    """
    sisters = others_in_sector(ticker, info["sector"], info["industry"])
    i = 0
    new_list = []
    while i < n and sisters:
        sister_ret = [create_dataframe(sisters[0], x)[0] for x in ["BS", "IS", "CF"]]
        blank = [x.empty for x in sister_ret]
        if True not in blank:
            vals = [sisters[0], sister_ret]
            new_list.append(vals)
            i += 1
        sisters.pop(0)
    return new_list


def clean_dataframes(*args) -> List[pd.DataFrame]:
    """
    All dataframes in the list take on the length of the shortest dataframe

    Parameters
    ----------
    *args : List[pd.DataFrame]
        List of dataframes to clean

    Returns
    -------
    dfs : List[pd.DataFrame]
        Cleaned list of dataframes
    """
    min_cols = min(x.shape[1] for x in args)
    dfs = [x.iloc[:, -min_cols:] for x in args]

    return dfs


def get_value(df: pd.DataFrame, row: str, column: int) -> Tuple[float, float]:
    """
    Gets a specific value from the dataframe

    Parameters
    ----------
    df : pd.DataFrame
        The dataframe to get the information from
    row : str
        The row to get the information from
    column : int
        The column to get the information from

    Returns
    -------
    value : List[float]
        The information in float format
    """
    val1: str = df.at[row, df.columns[column]]
    fin_val1: float = float(val1.replace(",", "").replace("-", "-0"))
    val2: str = df.at[row, df.columns[column + 1]]
    fin_val2: float = float(val2.replace(",", "").replace("-", "-0"))
    return fin_val1, fin_val2
